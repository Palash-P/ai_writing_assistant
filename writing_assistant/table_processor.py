# writing_assistant/table_processor.py
"""
Extract structured data from Excel and CSV files.
Converts tabular data into text that can be semantically searched.
"""
import logging

logger = logging.getLogger(__name__)


def excel_to_chunks(file_path, chunk_rows=20):
    """
    Convert Excel file to searchable text chunks.

    Why chunk by rows?
    - A 1000-row spreadsheet can't be sent to AI all at once
    - Grouping rows keeps related data together
    - Each chunk describes a section of the spreadsheet

    chunk_rows=20 means each chunk contains 20 rows of data.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    all_chunks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # First row is usually headers
        headers = [str(cell) if cell is not None else '' for cell in rows[0]]
        data_rows = rows[1:]

        logger.info(f"Sheet '{sheet_name}': {len(data_rows)} rows, {len(headers)} columns")

        # Process in chunks of chunk_rows rows
        for chunk_start in range(0, len(data_rows), chunk_rows):
            chunk_data = data_rows[chunk_start:chunk_start + chunk_rows]

            # Convert rows to readable text
            text_lines = [
                f"Sheet: {sheet_name} | Rows {chunk_start + 2}-{chunk_start + len(chunk_data) + 1}",
                f"Columns: {', '.join(h for h in headers if h)}",
                ""
            ]

            for row in chunk_data:
                # Create "Header: Value" pairs for each row
                row_parts = []
                for header, value in zip(headers, row):
                    if value is not None and str(value).strip():
                        row_parts.append(f"{header}: {value}")
                if row_parts:
                    text_lines.append(" | ".join(row_parts))

            chunk_text = "\n".join(text_lines)

            if chunk_text.strip():
                all_chunks.append({
                    'text': chunk_text,
                    'metadata': {
                        'page': 1,
                        'sheet': sheet_name,
                        'row_start': chunk_start + 2,
                        'row_end': chunk_start + len(chunk_data) + 1,
                        'chunk_type': 'spreadsheet',
                        'content_type': 'structured_data',
                    }
                })

    wb.close()
    logger.info(f"Excel → {len(all_chunks)} chunks")
    return all_chunks


def csv_to_chunks(file_path, chunk_rows=20):
    """Convert CSV to searchable text chunks"""
    import csv

    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return []

    headers = rows[0]
    data_rows = rows[1:]

    all_chunks = []

    for chunk_start in range(0, len(data_rows), chunk_rows):
        chunk_data = data_rows[chunk_start:chunk_start + chunk_rows]

        text_lines = [
            f"CSV Data | Rows {chunk_start + 2}-{chunk_start + len(chunk_data) + 1}",
            f"Columns: {', '.join(headers)}",
            ""
        ]

        for row in chunk_data:
            row_parts = [
                f"{header}: {value}"
                for header, value in zip(headers, row)
                if value.strip()
            ]
            if row_parts:
                text_lines.append(" | ".join(row_parts))

        chunk_text = "\n".join(text_lines)
        if chunk_text.strip():
            all_chunks.append({
                'text': chunk_text,
                'metadata': {
                    'page': 1,
                    'row_start': chunk_start + 2,
                    'row_end': chunk_start + len(chunk_data) + 1,
                    'chunk_type': 'csv',
                    'content_type': 'structured_data',
                }
            })

    logger.info(f"CSV → {len(all_chunks)} chunks")
    return all_chunks